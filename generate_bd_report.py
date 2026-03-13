#!/usr/bin/env python3
"""
商务部周报自动生成脚本
用法: python3 generate_bd_report.py --start 2026-03-06 --end 2026-03-12
"""

import argparse
import json
import os
import sys
import subprocess
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl 库，请运行: pip3 install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
CONFIG_FILE = SCRIPT_DIR / "config.json"
CUMULATIVE_FILE = SCRIPT_DIR / "cumulative.json"
MONTHLY_FILE = SCRIPT_DIR / "monthly.json"
REPORTS_DIR = SCRIPT_DIR / "reports"

SCORE_CATEGORIES = ["普通卡销售", "客户消费", "KOL销售", "白标卡销售", "API对接费", "卡面设计", "绑卡销售"]


def load_config():
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def load_cumulative():
    with open(CUMULATIVE_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_cumulative(data):
    with open(CUMULATIVE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_monthly():
    with open(MONTHLY_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_monthly(data):
    with open(MONTHLY_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def format_period(start_date, end_date):
    """Generate period string like '3.6-3.12' for file matching."""
    return f"{start_date.month}.{start_date.day}-{end_date.month}.{end_date.day}"


def find_source_file(source_dir, pattern, period):
    """Find source data file with flexible period matching."""
    # Try exact match first
    exact = os.path.join(source_dir, pattern.replace("{period}", period))
    if os.path.exists(exact):
        return exact
    # Try with leading zeros
    parts = period.split("-")
    if len(parts) == 2:
        # Try various date formats
        for f in os.listdir(source_dir):
            base = pattern.replace("{period}", "")
            if f.startswith(base.replace(".xlsx", "")) and f.endswith(".xlsx"):
                return os.path.join(source_dir, f)
    return None


def calc_consumer_spending(source_dir, period, config):
    """Calculate 客户消费积分 from 商务交易 file."""
    pattern = config["file_patterns"]["transaction"]
    filepath = find_source_file(source_dir, pattern, period)
    if not filepath:
        print(f"  ⚠ 未找到商务交易文件 (period={period})")
        return {bd: {"amount": 0, "score": 0, "users": 0, "txns": 0} for bd in config["bd_list"]}

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    email_map = config["email_mapping"]
    result = {bd: {"amount": 0.0, "score": 0.0, "users": set(), "txns": 0} for bd in config["bd_list"]}

    for row in ws.iter_rows(min_row=2, values_only=True):
        email = row[0]
        user_id = row[3]
        amount = float(row[4] or 0)
        score = float(row[5] or 0)
        bd = email_map.get(email)
        if bd and bd in result:
            result[bd]["amount"] += amount
            result[bd]["score"] += score
            result[bd]["txns"] += 1
            if user_id and amount > 0:
                result[bd]["users"].add(user_id)

    # Convert sets to counts
    for bd in result:
        result[bd]["users"] = len(result[bd]["users"])
    wb.close()

    rows = sum(1 for _ in ws.iter_rows(min_row=2))
    print(f"  ✓ 读取 {os.path.basename(filepath)} ({rows}条)")
    return result


def calc_card_sales(source_dir, period, config):
    """Calculate 普通卡销售积分 from 商务开卡 file."""
    pattern = config["file_patterns"]["card_opening"]
    filepath = find_source_file(source_dir, pattern, period)
    if not filepath:
        print(f"  ⚠ 未找到商务开卡文件 (period={period})")
        return {bd: {"cards": 0, "score": 0} for bd in config["bd_list"]}

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    email_map = config["email_mapping"]
    result = {bd: {"cards": 0, "score": 0.0} for bd in config["bd_list"]}

    rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows += 1
        email = row[0]
        card_type = str(row[5] or "")
        score = 0
        try:
            score = float(row[7] or 0)
        except (ValueError, TypeError):
            pass
        bd = email_map.get(email)
        if bd and bd in result and card_type != "无" and card_type:
            result[bd]["cards"] += 1
            result[bd]["score"] += score

    wb.close()
    print(f"  ✓ 读取 {os.path.basename(filepath)} ({rows}条)")
    return result


def calc_kol_sales(source_dir, period, config):
    """Calculate KOL销售积分 from KOL file + KOL-ID mapping."""
    # Load KOL-ID mapping
    special_file = os.path.join(source_dir, config["special_record_file"])
    if not os.path.exists(special_file):
        print(f"  ⚠ 未找到 {config['special_record_file']}")
        return {bd: {"total_score": 0, "phy_cards": 0, "vir_cards": 0, "users": 0, "cards": 0} for bd in config["bd_list"]}

    wb_special = openpyxl.load_workbook(special_file)
    ws_kol_id = wb_special[config["kol_id_sheet"]]
    kol_map = {}
    for row in ws_kol_id.iter_rows(min_row=2, values_only=True):
        kol_id = str(row[0]).strip() if row[0] else ""
        bd = str(row[1]).strip() if row[1] else ""
        if kol_id and bd:
            kol_map[kol_id] = bd
    wb_special.close()

    # Load KOL data
    pattern = config["file_patterns"]["kol"]
    filepath = find_source_file(source_dir, pattern, period)
    if not filepath:
        print(f"  ⚠ 未找到KOL文件 (period={period})")
        return {bd: {"total_score": 0, "phy_cards": 0, "vir_cards": 0, "users": 0, "cards": 0} for bd in config["bd_list"]}

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    result = {bd: {"total_score": 0, "phy_cards": 0, "vir_cards": 0, "users": set(), "cards": 0} for bd in config["bd_list"]}

    rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows += 1
        kol_id = str(row[1]).strip() if row[1] else ""
        total_score = int(row[4] or 0)
        phy_score = int(row[5] or 0)
        vir_score = int(row[6] or 0)
        user_id = row[3]
        bd = kol_map.get(kol_id)
        if bd and bd in result:
            result[bd]["total_score"] += total_score
            if phy_score > 0:
                result[bd]["phy_cards"] += 1
            if vir_score > 0:
                result[bd]["vir_cards"] += 1
            result[bd]["cards"] += (1 if phy_score > 0 or vir_score > 0 else 0)
            if user_id:
                result[bd]["users"].add(user_id)

    for bd in result:
        result[bd]["users"] = len(result[bd]["users"])
    wb.close()
    print(f"  ✓ 读取 {os.path.basename(filepath)} ({rows}条)")
    return result


def calc_special_scores(source_dir, config, start_date, end_date, mode="week"):
    """Calculate 白标卡/API/卡面设计/绑卡 from 特殊计分记录.

    mode="week": only records within [start_date, end_date]
    mode="cumulative": all records up to end_date (for historical total)
    mode="monthly": all records in end_date's month up to end_date
    """
    special_file = os.path.join(source_dir, config["special_record_file"])
    if not os.path.exists(special_file):
        return {bd: {"白标卡销售": 0, "API对接费": 0, "卡面设计": 0, "绑卡销售": 0} for bd in config["bd_list"]}

    wb = openpyxl.load_workbook(special_file, data_only=True)
    result = {bd: {"白标卡销售": 0, "API对接费": 0, "卡面设计": 0, "绑卡销售": 0} for bd in config["bd_list"]}

    # Determine which month sheets to scan
    if mode == "cumulative":
        months_to_check = list(range(1, end_date.month + 1))
    elif mode == "monthly":
        months_to_check = [end_date.month]
    else:
        months_to_check = list(set([start_date.month, end_date.month]))

    category_map = {
        "白标卡": "白标卡销售",
        "API": "API对接费",
        "API+定制": "API对接费",
        "卡面设计": "卡面设计",
        "绑卡": "绑卡销售",
    }

    sheets_read = []
    for month in months_to_check:
        month_sheet_name = f"{month} 月"
        if month_sheet_name not in wb.sheetnames:
            month_sheet_name = f"{month}月"
        if month_sheet_name not in wb.sheetnames:
            continue

        ws = wb[month_sheet_name]
        sheets_read.append(month_sheet_name)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            bd = str(row[0]).strip()
            date_str = str(row[1]).strip() if row[1] else ""
            category = str(row[2]).strip() if row[2] else ""
            score = row[5]

            # Try to get numeric score; handle formula cells
            try:
                score = float(score) if score is not None else 0
            except (ValueError, TypeError):
                score_str = str(score)
                if score_str.startswith("=") and "*" in score_str:
                    try:
                        quantity = float(row[4] or 0)
                        multiplier = float(score_str.split("*")[0].replace("=", ""))
                        score = multiplier * quantity
                    except (ValueError, IndexError):
                        score = 0
                else:
                    score = 0

            if date_str and bd in result:
                mapped_cat = category_map.get(category)
                if mapped_cat:
                    try:
                        parts = date_str.split(".")
                        d_month = int(parts[0])
                        d_day = int(parts[1])
                        d = datetime(start_date.year, d_month, d_day)
                        # Apply date filter based on mode
                        if mode == "week":
                            if start_date <= d <= end_date:
                                result[bd][mapped_cat] += score
                        elif mode == "cumulative":
                            if d <= end_date:
                                result[bd][mapped_cat] += score
                        elif mode == "monthly":
                            if d.month == end_date.month and d <= end_date:
                                result[bd][mapped_cat] += score
                    except (ValueError, IndexError):
                        result[bd][mapped_cat] += score

    wb.close()
    if sheets_read:
        print(f"  ✓ 读取 特殊计分记录/{', '.join(sheets_read)}")
    else:
        print(f"  ⚠ 未找到对应月份Sheet")
    return result


def fmt(n, decimals=None):
    """Format number with thousand separators."""
    if decimals is not None:
        n = round(n, decimals)
    if isinstance(n, float):
        if n == int(n) and decimals is None:
            return f"{int(n):,}"
        d = decimals if decimals is not None else 2
        return f"{n:,.{d}f}"
    return f"{n:,}"


def generate_html(weekly_scores, monthly_scores, cumulative, kol_data, config, start_date, end_date):
    """Generate the beautiful HTML report."""
    bd_list = config["bd_list"]
    badge = config["badge_classes"]

    start_str = f"{start_date.year}年{start_date.month}月{start_date.day}日"
    end_str = f"{end_date.year}年{end_date.month}月{end_date.day}日"
    period_short = f"{start_date.month}.{start_date.day} - {end_date.month}.{end_date.day}"
    period_dot = f"{start_date.year}.{start_date.month:02d}.{start_date.day:02d} — {end_date.year}.{end_date.month:02d}.{end_date.day:02d}"

    def score_row(bd, scores, total_label="周积分合计"):
        total = sum(scores.get(cat, 0) for cat in SCORE_CATEGORIES)
        return f"""        <tr>
          <td><span class="name-badge {badge[bd]}">{bd}</span></td>
          <td class="highlight-total">{fmt(total, 2)}</td>
          <td>{fmt(scores.get('普通卡销售', 0))}</td><td>{fmt(scores.get('客户消费', 0), 2)}</td><td>{fmt(scores.get('KOL销售', 0))}</td><td>{fmt(scores.get('白标卡销售', 0))}</td><td>{fmt(scores.get('API对接费', 0))}</td><td>{fmt(scores.get('绑卡销售', 0))}</td>
        </tr>"""

    def cumul_row(bd, scores):
        total = sum(scores.get(cat, 0) for cat in SCORE_CATEGORIES)
        return f"""        <tr>
          <td><span class="name-badge {badge[bd]}">{bd}</span></td>
          <td class="highlight-total">{fmt(total, 2)}</td>
          <td>{fmt(scores.get('普通卡销售', 0))}</td><td>{fmt(scores.get('客户消费', 0), 2)}</td><td>{fmt(scores.get('KOL销售', 0))}</td><td>{fmt(scores.get('白标卡销售', 0))}</td><td>{fmt(scores.get('API对接费', 0))}</td><td>{fmt(scores.get('绑卡销售', 0))}</td>
        </tr>"""

    weekly_rows = "\n".join(score_row(bd, weekly_scores[bd]) for bd in bd_list)
    monthly_rows = "\n".join(score_row(bd, monthly_scores[bd]) for bd in bd_list)
    cumul_rows = "\n".join(cumul_row(bd, cumulative[bd]) for bd in bd_list)

    # KOL section
    kol_stat_cards = ""
    kol_table_rows = ""
    for bd in bd_list:
        k = kol_data[bd]
        kol_stat_cards += f"""      <div class="stat-card">
        <div class="stat-name">{bd}</div>
        <div class="stat-value">{k['cards']}</div>
        <div class="stat-label">本周开卡数 | KOL 积分 {k['total_score']}</div>
      </div>\n"""
        kol_table_rows += f"""        <tr>
          <td><span class="name-badge {badge[bd]}">{bd}</span></td>
          <td>{k['users']}</td><td class="highlight">{k['cards']}</td><td>{k['phy_cards']}</td><td>{k['vir_cards']}</td><td class="highlight">{k['total_score']}</td>
        </tr>\n"""

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>商务部周数据报告 {period_dot}</title>
<style>
  @page {{ size: A4; margin: 20mm 15mm; }}
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: -apple-system, "PingFang SC", "Microsoft YaHei", "Helvetica Neue", sans-serif;
    color: #1a1a2e;
    background: #f0f2f5;
    line-height: 1.6;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}
  .container {{ max-width: 960px; margin: 0 auto; padding: 40px 20px; }}
  @media print {{
    body {{ background: #fff; }}
    .container {{ padding: 0; max-width: 100%; }}
    .section {{ break-inside: avoid; }}
  }}
  .header {{
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    color: #fff; padding: 40px 48px; border-radius: 16px;
    margin-bottom: 28px; position: relative; overflow: hidden;
  }}
  .header::after {{
    content: ''; position: absolute; top: -50%; right: -20%;
    width: 400px; height: 400px;
    background: radial-gradient(circle, rgba(83,178,235,0.15) 0%, transparent 70%);
    border-radius: 50%;
  }}
  .header h1 {{ font-size: 28px; font-weight: 700; letter-spacing: 2px; margin-bottom: 8px; position: relative; z-index: 1; }}
  .header .subtitle {{ font-size: 15px; color: rgba(255,255,255,0.7); font-weight: 400; position: relative; z-index: 1; }}
  .section {{
    background: #fff; border-radius: 12px; padding: 32px 36px; margin-bottom: 24px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.06), 0 4px 12px rgba(0,0,0,0.04);
  }}
  .section-title {{
    font-size: 18px; font-weight: 700; color: #1a1a2e;
    margin-bottom: 24px; padding-bottom: 12px; border-bottom: 2px solid #e8ecf1;
    display: flex; align-items: center; gap: 10px;
  }}
  .section-title .icon {{
    display: inline-flex; align-items: center; justify-content: center;
    width: 32px; height: 32px; border-radius: 8px; font-size: 16px; flex-shrink: 0;
  }}
  .icon-blue {{ background: #e8f4fd; color: #1976d2; }}
  .icon-green {{ background: #e8f5e9; color: #2e7d32; }}
  .sub-title {{ font-size: 14px; font-weight: 600; color: #546e7a; margin: 20px 0 12px; text-transform: uppercase; letter-spacing: 1px; }}
  table {{ width: 100%; border-collapse: separate; border-spacing: 0; font-size: 13px; margin-bottom: 16px; }}
  thead th {{
    background: #f5f7fa; color: #455a64; font-weight: 600; padding: 10px 14px;
    text-align: right; white-space: nowrap; border-bottom: 2px solid #dee2e6;
    font-size: 12px; text-transform: uppercase; letter-spacing: 0.5px;
  }}
  thead th:first-child {{ text-align: center; border-radius: 8px 0 0 0; }}
  thead th:last-child {{ border-radius: 0 8px 0 0; }}
  tbody td {{ padding: 10px 14px; text-align: right; border-bottom: 1px solid #f0f0f0; color: #37474f; }}
  tbody td:first-child {{ text-align: center; font-weight: 600; color: #1a1a2e; }}
  tbody tr:hover {{ background: #fafbfc; }}
  tbody tr:last-child td {{ border-bottom: none; }}
  .highlight {{ font-weight: 700; color: #1565c0; }}
  .highlight-total {{ font-weight: 700; color: #0d47a1; font-size: 14px; }}
  .name-badge {{ display: inline-block; padding: 2px 12px; border-radius: 20px; font-size: 12px; font-weight: 600; letter-spacing: 0.5px; }}
  .badge-captain {{ background: #fff3e0; color: #e65100; }}
  .badge-eliz {{ background: #e8f5e9; color: #2e7d32; }}
  .badge-if {{ background: #e3f2fd; color: #1565c0; }}
  .stats-row {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; margin-bottom: 20px; }}
  .stat-card {{
    background: linear-gradient(135deg, #f8f9fa, #fff);
    border: 1px solid #e8ecf1; border-radius: 10px; padding: 18px 20px; text-align: center;
  }}
  .stat-card .stat-name {{ font-size: 12px; color: #78909c; font-weight: 500; margin-bottom: 4px; text-transform: uppercase; letter-spacing: 0.5px; }}
  .stat-card .stat-value {{ font-size: 24px; font-weight: 700; color: #1a1a2e; }}
  .stat-card .stat-label {{ font-size: 11px; color: #90a4ae; margin-top: 2px; }}
  .footer {{ text-align: center; padding: 20px; color: #b0bec5; font-size: 12px; }}
</style>
</head>
<body>
<div class="container">

  <div class="header">
    <h1>商务部周数据报告</h1>
    <div class="subtitle">{start_str} — {end_str}</div>
  </div>

  <div class="section">
    <div class="section-title">
      <span class="icon icon-blue">📊</span>
      一、商务团队积分统计
    </div>

    <div class="sub-title">1.1 本周积分（{period_short}）</div>
    <table>
      <thead>
        <tr>
          <th style="text-align:center">商务经理</th>
          <th>周积分合计</th><th>普通卡销售</th><th>客户消费</th><th>KOL 销售</th><th>白标卡销售</th><th>API 对接费</th><th>绑卡销售</th>
        </tr>
      </thead>
      <tbody>
{weekly_rows}
      </tbody>
    </table>

    <div class="sub-title">1.2 本月累计积分（{start_date.month}月）</div>
    <table>
      <thead>
        <tr>
          <th style="text-align:center">商务经理</th>
          <th>月累计积分</th><th>普通卡销售</th><th>客户消费</th><th>KOL 销售</th><th>白标卡销售</th><th>API 对接费</th><th>绑卡销售</th>
        </tr>
      </thead>
      <tbody>
{monthly_rows}
      </tbody>
    </table>

    <div class="sub-title">1.3 历史累计积分（截至 {end_date.month}.{end_date.day}）</div>
    <table>
      <thead>
        <tr>
          <th style="text-align:center">商务经理</th>
          <th>累计总积分</th><th>普通卡销售</th><th>客户消费</th><th>KOL 销售</th><th>白标卡销售</th><th>API 对接费</th><th>绑卡销售</th>
        </tr>
      </thead>
      <tbody>
{cumul_rows}
      </tbody>
    </table>
  </div>

  <div class="section">
    <div class="section-title">
      <span class="icon icon-green">📈</span>
      二、本周商务部 KOL 贡献（{period_short}）
    </div>

    <div class="stats-row">
{kol_stat_cards}    </div>

    <table>
      <thead>
        <tr>
          <th style="text-align:center">商务经理</th>
          <th>本周开卡用户数</th><th>本周开卡数</th><th>实体卡</th><th>虚拟卡</th><th>KOL 积分</th>
        </tr>
      </thead>
      <tbody>
{kol_table_rows}      </tbody>
    </table>
  </div>

  <div class="footer">
    本报告由数据自动化系统生成 &nbsp;|&nbsp; 数据周期：{period_dot}
  </div>

</div>
</body>
</html>"""
    return html


def generate_md(weekly_scores, monthly_scores, cumulative, kol_data, config, start_date, end_date):
    """Generate Markdown report."""
    bd_list = config["bd_list"]
    period_short = f"{start_date.month}.{start_date.day} - {end_date.month}.{end_date.day}"
    start_str = f"{start_date.year}年{start_date.month}月{start_date.day}日"
    end_str = f"{end_date.year}年{end_date.month}月{end_date.day}日"
    period_dot = f"{start_date.year}.{start_date.month:02d}.{start_date.day:02d} — {end_date.year}.{end_date.month:02d}.{end_date.day:02d}"

    def score_line(bd, scores):
        total = sum(scores.get(cat, 0) for cat in SCORE_CATEGORIES)
        return f"| **{bd}** | **{fmt(total, 2)}** | {fmt(scores.get('普通卡销售', 0))} | {fmt(scores.get('客户消费', 0), 2)} | {fmt(scores.get('KOL销售', 0))} | {fmt(scores.get('白标卡销售', 0))} | {fmt(scores.get('API对接费', 0))} | {fmt(scores.get('绑卡销售', 0))} |"

    lines = [
        f"# 商务部周数据报告",
        f"**{start_str} — {end_str}**",
        "",
        "---",
        "",
        "## 一、商务团队积分统计",
        "",
        f"### 1.1 本周积分（{period_short}）",
        "",
        "| 商务经理 | 周积分合计 | 普通卡销售 | 客户消费 | KOL 销售 | 白标卡销售 | API 对接费 | 绑卡销售 |",
        "|:--------:|----------:|----------:|---------:|---------:|----------:|----------:|---------:|",
    ]
    for bd in bd_list:
        lines.append(score_line(bd, weekly_scores[bd]))

    lines += [
        "",
        f"### 1.2 本月累计积分（{start_date.month}月）",
        "",
        "| 商务经理 | 月累计积分 | 普通卡销售 | 客户消费 | KOL 销售 | 白标卡销售 | API 对接费 | 绑卡销售 |",
        "|:--------:|----------:|----------:|---------:|---------:|----------:|----------:|---------:|",
    ]
    for bd in bd_list:
        lines.append(score_line(bd, monthly_scores[bd]))

    lines += [
        "",
        f"### 1.3 历史累计积分（截至 {end_date.month}.{end_date.day}）",
        "",
        "| 商务经理 | 累计总积分 | 普通卡销售 | 客户消费 | KOL 销售 | 白标卡销售 | API 对接费 | 绑卡销售 |",
        "|:--------:|-----------:|----------:|---------:|---------:|----------:|----------:|---------:|",
    ]
    for bd in bd_list:
        lines.append(score_line(bd, cumulative[bd]))

    lines += [
        "",
        "---",
        "",
        f"## 二、本周商务部 KOL 贡献（{period_short}）",
        "",
        "| 商务经理 | 本周开卡用户数 | 本周开卡数 | 实体卡 | 虚拟卡 | KOL 积分 |",
        "|:--------:|----------:|--------:|------:|------:|--------:|",
    ]
    for bd in bd_list:
        k = kol_data[bd]
        lines.append(f"| **{bd}** | {k['users']} | {k['cards']} | {k['phy_cards']} | {k['vir_cards']} | {k['total_score']} |")

    lines += [
        "",
        "---",
        "",
        f"> 本报告由数据自动化系统生成 | 数据周期：{period_dot}",
    ]

    return "\n".join(lines)


def generate_index_html(reports):
    """Generate index.html with list of all reports."""
    reports_sorted = sorted(reports, key=lambda r: r["start"], reverse=True)
    latest = reports_sorted[0] if reports_sorted else None

    report_links = ""
    for r in reports_sorted:
        report_links += f'      <a href="{r["path"]}" class="report-link">\n'
        report_links += f'        <span class="date">{r["label"]}</span>\n'
        if r == latest:
            report_links += f'        <span class="badge-latest">最新</span>\n'
        report_links += f'      </a>\n'

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Tevau 商务部周报</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif;
    background: #f0f2f5; color: #1a1a2e; min-height: 100vh;
    display: flex; align-items: center; justify-content: center;
  }}
  .container {{ max-width: 480px; width: 100%; padding: 40px 20px; }}
  .header {{
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    color: #fff; padding: 36px; border-radius: 16px; margin-bottom: 24px; text-align: center;
  }}
  .header h1 {{ font-size: 22px; font-weight: 700; letter-spacing: 2px; margin-bottom: 6px; }}
  .header p {{ font-size: 13px; color: rgba(255,255,255,0.6); }}
  .reports {{
    background: #fff; border-radius: 12px; overflow: hidden;
    box-shadow: 0 1px 3px rgba(0,0,0,0.06), 0 4px 12px rgba(0,0,0,0.04);
  }}
  .reports-title {{ padding: 16px 20px; font-size: 13px; color: #78909c; font-weight: 600; text-transform: uppercase; letter-spacing: 1px; border-bottom: 1px solid #f0f0f0; }}
  .report-link {{
    display: flex; align-items: center; justify-content: space-between;
    padding: 14px 20px; color: #1a1a2e; text-decoration: none;
    border-bottom: 1px solid #f5f5f5; transition: background 0.15s;
  }}
  .report-link:hover {{ background: #f8f9fa; }}
  .report-link:last-child {{ border-bottom: none; }}
  .report-link .date {{ font-size: 14px; font-weight: 500; }}
  .badge-latest {{
    font-size: 11px; background: #e8f5e9; color: #2e7d32;
    padding: 2px 10px; border-radius: 12px; font-weight: 600;
  }}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>商务部周报</h1>
    <p>Tevau Business Department Weekly Reports</p>
  </div>
  <div class="reports">
    <div class="reports-title">历史报告</div>
{report_links}  </div>
</div>
</body>
</html>"""


def update_index(reports_dir):
    """Scan reports directory and update index.html."""
    reports = []
    for year_dir in sorted(reports_dir.iterdir()):
        if year_dir.is_dir() and year_dir.name.isdigit():
            for f in sorted(year_dir.iterdir()):
                if f.suffix == ".html":
                    # Parse filename like 0227-0305.html
                    name = f.stem
                    parts = name.split("-")
                    if len(parts) == 2:
                        s, e = parts
                        label = f"{year_dir.name}.{s[:2]}.{s[2:]} — {year_dir.name}.{e[:2]}.{e[2:]}"
                        reports.append({
                            "path": f"reports/{year_dir.name}/{f.name}",
                            "label": label,
                            "start": f"{year_dir.name}{s}",
                        })

    index_html = generate_index_html(reports)
    index_path = reports_dir.parent / "index.html"
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(index_html)
    return index_path


def main():
    parser = argparse.ArgumentParser(description="生成商务部周数据报告")
    parser.add_argument("--start", required=True, help="周起始日期 (YYYY-MM-DD)")
    parser.add_argument("--end", required=True, help="周结束日期 (YYYY-MM-DD)")
    parser.add_argument("--push", action="store_true", help="自动推送到 GitHub")
    parser.add_argument("--source-dir", help="源数据目录 (覆盖 config.json)")
    args = parser.parse_args()

    start_date = datetime.strptime(args.start, "%Y-%m-%d")
    end_date = datetime.strptime(args.end, "%Y-%m-%d")

    print(f"\n{'='*50}")
    print(f"  商务部周报生成器")
    print(f"  周期: {start_date.strftime('%Y.%m.%d')} — {end_date.strftime('%Y.%m.%d')}")
    print(f"{'='*50}\n")

    # Load config
    config = load_config()
    source_dir = args.source_dir or config["source_data_dir"]
    period = format_period(start_date, end_date)

    print("📂 读取源数据...")

    # 1. Calculate all scores
    spending = calc_consumer_spending(source_dir, period, config)
    cards = calc_card_sales(source_dir, period, config)
    kol = calc_kol_sales(source_dir, period, config)
    special_week = calc_special_scores(source_dir, config, start_date, end_date, mode="week")
    special_month = calc_special_scores(source_dir, config, start_date, end_date, mode="monthly")

    # 2. Assemble weekly scores
    weekly_scores = {}
    for bd in config["bd_list"]:
        weekly_scores[bd] = {
            "普通卡销售": cards[bd]["score"],
            "客户消费": round(spending[bd]["score"], 2),
            "KOL销售": kol[bd]["total_score"],
            "白标卡销售": special_week[bd]["白标卡销售"],
            "API对接费": special_week[bd]["API对接费"],
            "卡面设计": special_week[bd]["卡面设计"],
            "绑卡销售": special_week[bd]["绑卡销售"],
        }

    # 3. Print weekly scores summary
    print(f"\n{'─'*40}")
    print("📊 本周积分:")
    for bd in config["bd_list"]:
        total = sum(weekly_scores[bd].values())
        print(f"  {bd:6s}: {total:>10,.2f} 分")
    print(f"{'─'*40}")

    # 4. Load cumulative and update (incremental for all categories)
    cumul_data = load_cumulative()
    SPECIAL_CATS = ["白标卡销售", "API对接费", "卡面设计", "绑卡销售"]
    cumulative = {}
    for bd in config["bd_list"]:
        prev = cumul_data["data"].get(bd, {cat: 0 for cat in SCORE_CATEGORIES})
        cumulative[bd] = {}
        for cat in SCORE_CATEGORIES:
            cumulative[bd][cat] = prev.get(cat, 0) + weekly_scores[bd].get(cat, 0)

    # 5. Monthly scores — load monthly.json and accumulate
    monthly_data = load_monthly()
    report_month = end_date.month
    report_year = end_date.year

    # If month changed, reset monthly accumulator
    if monthly_data["year"] != report_year or monthly_data["month"] != report_month:
        monthly_data = {
            "year": report_year,
            "month": report_month,
            "data": {bd: {cat: 0 for cat in SCORE_CATEGORIES} for bd in config["bd_list"]}
        }

    monthly_scores = {}
    for bd in config["bd_list"]:
        prev_month = monthly_data["data"].get(bd, {cat: 0 for cat in SCORE_CATEGORIES})
        monthly_scores[bd] = {}
        for cat in SCORE_CATEGORIES:
            if cat in SPECIAL_CATS:
                # For special cats, use full month scan (always accurate, catches retroactive adds)
                monthly_scores[bd][cat] = special_month[bd].get(cat, 0)
            else:
                # For regular cats, accumulate from previous weeks in this month
                monthly_scores[bd][cat] = prev_month.get(cat, 0) + weekly_scores[bd].get(cat, 0)

    # 6. KOL data for report
    kol_data = {bd: kol[bd] for bd in config["bd_list"]}

    # 7. Generate reports
    print("\n📝 生成报告...")

    year_dir = REPORTS_DIR / str(start_date.year)
    year_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{start_date.month:02d}{start_date.day:02d}-{end_date.month:02d}{end_date.day:02d}"

    # HTML
    html = generate_html(weekly_scores, monthly_scores, cumulative, kol_data, config, start_date, end_date)
    html_path = year_dir / f"{filename}.html"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✓ {html_path.relative_to(SCRIPT_DIR)}")

    # MD
    md = generate_md(weekly_scores, monthly_scores, cumulative, kol_data, config, start_date, end_date)
    md_path = year_dir / f"{filename}.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md)
    print(f"  ✓ {md_path.relative_to(SCRIPT_DIR)}")

    # Update index
    index_path = update_index(REPORTS_DIR)
    print(f"  ✓ index.html (首页已更新)")

    # 8. Update cumulative data
    cumul_data["last_updated"] = end_date.strftime("%Y-%m-%d")
    cumul_data["data"] = cumulative
    save_cumulative(cumul_data)
    print(f"  ✓ cumulative.json (历史累计已更新)")

    # 9. Update monthly data
    monthly_data["year"] = report_year
    monthly_data["month"] = report_month
    monthly_data["data"] = monthly_scores
    save_monthly(monthly_data)
    print(f"  ✓ monthly.json (月累计已更新)")

    # 9. Push to GitHub if requested
    if args.push:
        print("\n🚀 推送到 GitHub...")
        try:
            subprocess.run(["git", "add", "."], cwd=SCRIPT_DIR, check=True)
            msg = f"update: 商务部周报 {start_date.strftime('%m.%d')}-{end_date.strftime('%m.%d')}"
            subprocess.run(["git", "commit", "-m", msg], cwd=SCRIPT_DIR, check=True)
            subprocess.run(["git", "push"], cwd=SCRIPT_DIR, check=True)
            print("  ✓ 已推送到 GitHub Pages")
        except subprocess.CalledProcessError as e:
            print(f"  ✗ 推送失败: {e}")

    print(f"\n✅ 完成！报告文件: {html_path}")
    if args.push:
        print(f"🌐 在线地址即将更新")
    print()


if __name__ == "__main__":
    main()
